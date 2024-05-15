import openpyxl as xl
import re
import streamlit as st

def get_routing(vpn,sheet):
    i = 0
    list_vnf = []
    list_rt = []
    list_Merge = []
    list_t1 = []
    list_t2 = []
    list_t3 = []
    list_t4 = []
    for v in sheet.values:
        i+=1
        if v[2] == vpn:
            for r in sheet.iter_rows(max_row=i,min_row=i):
                for v in r:
                    if v.value != None:
                        list_vnf.append(v.value)

        elif v[1] == vpn:
            list_rt.append(i)


    for x in list_rt:
        z = x+1
        list_Merge.append(z)

    for c in sheet.iter_rows(min_row=list_rt[0],max_row=list_rt[0]):
        for v in c:
            list_t1.append(v.value)

    for c in sheet.iter_rows(min_row=list_rt[1],max_row=list_rt[1]):
        for v in c:
            if v.value != None:
                list_t2.append(v.value)

    for c in sheet.iter_rows(min_row=list_rt[2],max_row=list_Merge[2]):
        for v in c:
            if v.value != None:
                list_t3.append(v.value)

    for c in sheet.iter_rows(min_row=list_rt[3],max_row=list_Merge[3]):
        for v in c:
            if v.value != None:
                list_t4.append(v.value)

#print(list_vnf)
    return list_t1,list_t2,list_t3,list_t4


# print(list_vnf,list_rt[0:25],list_rt[25:36])
def get_intIp(vpn,sheet):
    #sh_intIp = wb['Internal IP Assignment']
    list_intIp = []
    i = 0
    for v in sheet.values:
        i += 1
        if v[2] == vpn:
            for r in sheet.iter_rows(min_row=i,max_row=i):
                for v in r:
                    if v.value != None:
                        list_intIp.append(v.value)

    return list_intIp

#
# class WorkSheet:
#     def __init__(self,vpn,sh_routing,sh_intIp):
#         self.vpn = vpn
#         self.sh_routing = sh_routing
#         self.sh_intIp = sh_intIp
#
#     def get_routing(self,vpn,sh_routing):
#         i = 0
#         list_vnf = []
#         list_rt = []
#         list_Merge = []
#         list_t1 = []
#         list_t2 = []
#         list_t3 = []
#         list_t4 = []
#         for v in sh_routing.values:
#             i+=1
#             if v[2] == vpn:
#                 for r in sh_routing.iter_rows(max_row=i,min_row=i):
#                     for v in r:
#                         if v.value != None:
#                             list_vnf.append(v.value)
#
#             elif v[1] == vpn:
#                 list_rt.append(i)
#
#         for x in list_rt:
#             z = x+1
#             list_Merge.append(z)
#
#         for c in sh_routing.iter_rows(min_row=list_rt[0],max_row=list_rt[0]):
#             for v in c:
#                 list_t1.append(v.value)
#
#         for c in sh_routing.iter_rows(min_row=list_rt[1],max_row=list_rt[1]):
#             for v in c:
#                 if v.value != None:
#                     list_t2.append(v.value)
#
#         for c in sh_routing.iter_rows(min_row=list_rt[2],max_row=list_Merge[2]):
#             for v in c:
#                 if v.value != None:
#                     list_t3.append(v.value)
#
#         for c in sh_routing.iter_rows(min_row=list_rt[3],max_row=list_Merge[3]):
#             for v in c:
#                 if v.value != None:
#                     list_t4.append(v.value)
#
#     #print(list_vnf)
#         return list_t1,list_t2,list_t3,list_t4
#
#
#     # print(list_vnf,list_rt[0:25],list_rt[25:36])
#     def get_intIp(self,vpn,sheet):
#         #sh_intIp = wb['Internal IP Assignment']
#         list_intIp = []
#         i = 0
#         for v in sheet.values:
#             i += 1
#             if v[2] == vpn:
#                 for r in sheet.iter_rows(min_row=i,max_row=i):
#                     for v in r:
#                         if v.value != None:
#                             list_intIp.append(v.value)
#
#         return list_intIp
#
#     def test(self,vpn,sheet):
#         i = 0
#         list_vnf = []
#         list_rt = []
#         list_Merge = []
#         list_t1 = []
#         list_t2 = []
#         list_t3 = []
#         list_t4 = []
#         for v in sheet.values:
#             b = v[i][i]
#             i += 1
#             return b
