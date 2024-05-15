import glob
from datetime import datetime
# from script_sla_st2 import *
# from script_tls_st2 import *
import openpyxl as xl
import time
import sys
import os
import streamlit as st
st.set_page_config(layout="wide",page_title='Gen WAN Script',page_icon=':black_small_square:')
st.header("Generate WAN script")


class SLA_Workbook:
    def __init__(self,vpn,workbook):
        self.vpn = vpn
        self.workbook = workbook


    def get_ipAssingment(self,vpn,sheet):
        i = 0
        list_rowid = []
        list_ip_dcgw01 = []
        list_ip_dcgw02 = []
        for c in sheet.values:
            i += 1
            if c[6] == vpn:
                list_rowid.append(i)

        for r in sheet.iter_rows(min_row=list_rowid[0], max_row=list_rowid[0]):
            for v in r:
                list_ip_dcgw01.append(v.value)

        for r in sheet.iter_rows(min_row=list_rowid[1], max_row=list_rowid[1]):
            for v in r:
                list_ip_dcgw02.append(v.value)
        return list_ip_dcgw01,list_ip_dcgw02

    def get_intIp(self,vpn,sheet):
        i = 0
        list_rowid = []
        list_intIp_dcgw01 = []
        list_intIp_dcgw02 = []
        for c in sheet.values:
            i += 1
            if c[2] == vpn:
                list_rowid.append(i)

        for r in sheet.iter_rows(min_row=list_rowid[0], max_row=list_rowid[0]):
            for v in r:
                list_intIp_dcgw01.append(v.value)

        for r in sheet.iter_rows(min_row=list_rowid[1], max_row=list_rowid[1]):
            for v in r:
                list_intIp_dcgw02.append(v.value)
        return list_intIp_dcgw01,list_intIp_dcgw02



    def get_route(self,sheet,vpn):
          #get value for formular cell
          list1_vpn_rowid = []
          list2_vpn_rowid = []
          list3_vpn_rowid = []
          i = 0
          for v in sheet.values:
              i += 1
              if v[1] == vpn:
                  list1_vpn_rowid.append(i)

              if v[2] == vpn:
                  z = i + 1
                  list3_vpn_rowid.append(i)
                  list3_vpn_rowid.insert(1, z)
          # print(list3_vpn_rowid)

          for x in list1_vpn_rowid:
              n = x + 1
              list2_vpn_rowid.append(n)

          list1_vpn_rowid.insert(1, list2_vpn_rowid[0])
          list1_vpn_rowid.insert(3, list2_vpn_rowid[1])
          # print(list1_vpn_rowid)

          list_dcgw01_t1 = []
          for t1 in list3_vpn_rowid:
              for c in sheet.iter_rows(min_row=t1, max_row=t1):
                  for z in c:
                      if z.value != None:
                        list_dcgw01_t1.append(z.value)
          # print(list_dcgw01_t1)

          list_dcgw01 = []
          for l in list1_vpn_rowid:
              for v in sheet.iter_rows(min_row=l, max_row=l):
                  for y in v:
                      if y.value != None:
                          list_dcgw01.append(y.value)
          # print(list_dcgw01)

          return list_dcgw01_t1,list_dcgw01




###################NFVI-TLS####################################################

class TLS_Workbook:
    def __init__(self, vpn, workbook):
        self.vpn = vpn
        self.workbook = workbook

    def get_routing(self,vpn,sheet):
        i = 0
        list_vnf = []
        list_rt = []
        list_Merge = []
        list_t1 = []
        list_t2 = []
        list_t3 = []
        list_t4 = []
        for v in sheet.values:
            i+=1
            if v[2] == vpn:
                for r in sheet.iter_rows(max_row=i,min_row=i):
                    for v in r:
                        if v.value != None:
                            list_vnf.append(v.value)

            elif v[1] == vpn:
                list_rt.append(i)


        for x in list_rt:
            z = x+1
            list_Merge.append(z)

        for c in sheet.iter_rows(min_row=list_rt[0],max_row=list_rt[0]):
            for v in c:
                list_t1.append(v.value)

        for c in sheet.iter_rows(min_row=list_rt[1],max_row=list_rt[1]):
            for v in c:
                if v.value != None:
                    list_t2.append(v.value)

        for c in sheet.iter_rows(min_row=list_rt[2],max_row=list_Merge[2]):
            for v in c:
                if v.value != None:
                    list_t3.append(v.value)

        for c in sheet.iter_rows(min_row=list_rt[3],max_row=list_Merge[3]):
            for v in c:
                if v.value != None:
                    list_t4.append(v.value)

    #print(list_vnf)
        return list_t1,list_t2,list_t3,list_t4


    # print(list_vnf,list_rt[0:25],list_rt[25:36])
    def get_intIp(self,vpn,sheet):
        #sh_intIp = wb['Internal IP Assignment']
        list_intIp = []
        i = 0
        for v in sheet.values:
            i += 1
            if v[2] == vpn:
                for r in sheet.iter_rows(min_row=i,max_row=i):
                    for v in r:
                        if v.value != None:
                            list_intIp.append(v.value)

        return list_intIp



if __name__ == '__main__':
    try:
        st.text_input("Enter VPN", key="name")
        uploaded_lld = st.file_uploader("Upload LLD", type=["xlsx"])
        wb = xl.load_workbook(uploaded_lld, data_only=True)
        vpn_name = st.session_state.name
        lld_file_name = uploaded_lld.name.split(' ')
        try:
            if lld_file_name[2] == 'DCTL22H':
                wb_tls = TLS_Workbook(vpn_name,wb)
                sh_routing = wb_tls.workbook['Routing']
                sh_intIp = wb_tls.workbook['Internal IP Assignment']
                tls_data = wb_tls.get_routing(vpn_name, sh_routing)
                inIp = wb_tls.get_intIp(vpn_name, sh_intIp)
                vrf_name = tls_data[2][0]
                vlan04 = tls_data[2][1].split('.')[-1]
                vlan05 = tls_data[2][6].split('.')[-1]
                upftls_vlan = inIp[1]
                rn04_eor01_ip = tls_data[2][2].split("/")[-2]
                rn05_eor01_ip = tls_data[2][7].split("/")[-2]
                ipcbb01_ip = tls_data[2][5].split("/")[-2]
                ipcbb02_ip = tls_data[2][10].split("/")[-2]
                process_id = tls_data[3][5]
                router_id = tls_data[0][18]
                as_number = tls_data[3][2]
                as_number_ipcbb = tls_data[3][8]
                eor_gwIp = tls_data[0][22]

                if len(tls_data[3]) == 20:
                    authentication_key01 = "peer {} password cipher {}".format(ipcbb01_ip, tls_data[3][19])
                    authentication_key02 = "peer {} password cipher {}".format(ipcbb02_ip, tls_data[3][19])
                else:
                    authentication_key = "#No BGP authentication key"
                # print(rn04_eor01_ip,rn05_eor01_ip)
                # print(ipcbb01_ip,ipcbb02_ip)
                # print(as_number,process_id,router_id)
                # print(eor_gwIp)
                # print(tls_data)

                # print(vlan04)
                # print(vrf_name)
                #print(tls_data[0],'\n',tls_data[1],'\n',tls_data[2],'\n',tls_data[3])

                NFVITL22H_RN04RN05 = ('''///////////NFVITL22H_RN04&RN05_EOR01 (Stack)///////////////
///create batch vlan
#
vlan batch {vlan04} {vlan05}
#
///Bind VLAN to Eth-trunk(Existing)
#
interface Eth-trunk39
port trunk allow-pass vlan {vlan04}
#
interface Eth-trunk40
port trunk allow-pass vlan {vlan05}
#
///create vpn instance
#
ip vpn-instance {vrf_name}
 ipv4-family
  route-distinguisher {rn04_eor01_ip}:100
#
///create interface Vlanif for WAN to AIS IPCBB
#
interface Vlanif{vlan04}
 description EOR_To_IPCBB_{vrf_name}
 ip binding vpn-instance {vrf_name}
 ip address {rn04_eor01_ip} 255.255.255.252
#
interface Vlanif{vlan05}
 description EOR_To_IPCBB_{vrf_name}
 ip binding vpn-instance {vrf_name}
 ip address {rn05_eor01_ip} 255.255.255.252
 
///Once you have reached this step You can test ping by following command.///
//ping -vpn-instance {vrf_name} -a {rn04_eor01_ip} {ipcbb01_ip}
//ping -vpn-instance {vrf_name} -a {rn05_eor01_ip} {ipcbb02_ip} 
#
////Peer eBGP between DCGW <-> AIS IPCBB
#
bgp 65504
ipv4-family vpn-instance {vrf_name}
 router-id {rn04_eor01_ip}
 import-route ospf {process_id} route-policy SGi_BGP_import
 maximum load-balancing ebgp 2
 peer {ipcbb01_ip} as-number {as_number_ipcbb}
 peer {ipcbb01_ip} connect-interface Vlanif{vlan04}
 peer {ipcbb01_ip} local-as {as_number}
 {authentication_key01}
 peer {ipcbb02_ip} as-number {as_number_ipcbb}
 peer {ipcbb02_ip} connect-interface Vlanif{vlan05}
 peer {ipcbb02_ip} local-as {as_number}
 {authentication_key02}

#

////create ospf between (DCGW<->UPF)
#
ospf {process_id} router-id {router_id} vpn-instance {vrf_name}
 description UPFTL21H_{vrf_name}
 default-route-advertise
 bfd all-interfaces enable
 bfd all-interfaces min-tx-interval 500 min-rx-interval 500 detect-multiplier 4
 import-route static
 import-route bgp
 vpn-instance-capability simple
 area 0.0.0.10
#

create interface Vlanif for ospf, ospf-bfd
#
interface Vlanif{upftls_vlan}
 description UPFTL21H_{vrf_name}
 ip binding vpn-instance {vrf_name}
 ip address {eor_gwIp} 255.255.255.248
 ospf bfd enable
 ospf bfd min-tx-interval 500 min-rx-interval 500 detect-multiplier 4
 ospf enable {process_id} area 0.0.0.10
#

commit
save

==== Command to Check eBGP state to IPCBB (Shall be Established) =====

display bgp vpnv4 vpn-instance {vrf_name} peer

===== Command to Check OSPF state to UPF (Shall be FULL) =====

display ospf {process_id} peer


===================Rollback=====================================

# undo interface Vlanif {upftls_vlan}
# undo ospf {process_id}
# undo ipv4-family vpn-instance {vrf_name}
# undo interface Vlanif {vlan04}
# undo interface Vlanif {vlan05}
# undo ip vpn-instance {vrf_name}

# 
interface Eth-trunk39
 undo port trunk allow-pass vlan {vlan04}
#
interface Eth-trunk40
 undo port trunk allow-pass vlan {vlan05}
 
# undo vlan {vlan04}
# undo vlan {vlan05}

===== Commit and Save =====
commit
q
save
''' .format(vrf_name=vrf_name,vlan04=vlan04,vlan05=vlan05,rn04_eor01_ip=rn04_eor01_ip,rn05_eor01_ip=rn05_eor01_ip,
                           ipcbb01_ip=ipcbb01_ip,ipcbb02_ip=ipcbb02_ip,process_id=process_id,as_number=as_number,as_number_ipcbb=as_number_ipcbb,
                           router_id=router_id,eor_gwIp=eor_gwIp,upftls_vlan=upftls_vlan,authentication_key01=authentication_key01,authentication_key02=authentication_key02))

                    # print(NFVITL22H_RN04RN05)
                col1, col2 = st.columns(2)
                created_file = datetime.today().strftime("%Y%m%d")
                file_name = f'./TLS/NFVI/NFVITL22H_RN04&RN05_EOR01-{vrf_name}-{created_file}.txt'
                with open(file_name,"w+") as f:
                    f.write(NFVITL22H_RN04RN05)
                with col1:
                    container = st.container(border=True)
                    with open(file_name, 'r') as file:
                        for text in file:
                            container.text(text)  # .text does not do any formatting
                            time.sleep(0.05)

        #/////////////////////////////////////////////////////////

                    upftls_routerid = tls_data[0][3]
                    upftls_intIp = tls_data[0]
                    upftls_ip_trunk1 = inIp[4]

                    list_upftls_ip_trunk = upftls_ip_trunk1.replace("'",'')
                    list_upftls_ip_trunk = list_upftls_ip_trunk.split('.')
                    ip_last_digit = int(list_upftls_ip_trunk[3]) + 1
                    list_upftls_ip_trunk[3] = str(ip_last_digit)
                    upftls_ip_trunk2 = '.'.join(list_upftls_ip_trunk)

                # print(inIp)
                # print(upftls_routerid)

                    UPFTL21H = '''///////////UPFTL21H///////////////
ADD L3VPNINST:VRFNAME="{vrf_name}";
ADD VPNINSTAF:VRFNAME="{vrf_name}",AFTYPE=ipv4uni,VRFRD="{upftls_ip_trunk1}:100",LOCALCROSSNHPMOD=FALSE,EXPLYADDERTFIRST=FALSE,VRFLABELMODE=perRoute,FRRENABLE=FALSE;
ADD INTERFACE:IFNAME="Eth-trunk1.{upftls_vlan}";
ADD INTERFACE:IFNAME="Eth-trunk2.{upftls_vlan}";
ADD IPBINDVPN:IFNAME="Eth-trunk1.{upftls_vlan}",VRFNAME="{vrf_name}";
ADD IPBINDVPN:IFNAME="Eth-trunk2.{upftls_vlan}",VRFNAME="{vrf_name}";
ADD ETHSUBIF:IFNAME="Eth-trunk1.{upftls_vlan}",VLANTYPEVID={upftls_vlan};
ADD ETHSUBIF:IFNAME="Eth-trunk2.{upftls_vlan}",VLANTYPEVID={upftls_vlan};
ADD IFIPV4ADDRESS:IFNAME="Eth-trunk1.{upftls_vlan}",IFIPADDR="{upftls_ip_trunk1}",SUBNETMASK="255.255.255.248";
ADD IFIPV4ADDRESS:IFNAME="Eth-trunk2.{upftls_vlan}",IFIPADDR="{upftls_ip_trunk2}",SUBNETMASK="255.255.255.248";
ADD OSPF:PROCID={process_id},VRFNAME="{vrf_name}",SCHEMAROUID="{upftls_routerid}",BFDALLINTFFLG=TRUE,BFDRXCFGFLAG=TRUE,BFDMINRXINTV=500,BFDTXCFGFLAG=TRUE,BFDMINTXINTV=500,SCHEMABFDDETMUL=4,LSAARRMAXINTV=1000,LSAARRSTARINTV=500,LSAARRHLDINTV=500,VPNINSCAPSIMFLG=TRUE,SPFMAXINTERVAL=5000,SPFSTARTINTV=50,SPFHOLDINTERVAL=200;
ADD OSPFAREA:PROCID={process_id},AREAID="0.0.0.10",AREATYPE=Normal,MAXCOSTFLAG=FALSE;
ADD OSPFINTERFACE:PROCID={process_id},AREAID="0.0.0.10",IFNAME="Eth-trunk1.{upftls_vlan}",NETWORKTYPE=broadcast,HOLDDOWNITVL=10,FRRBLOCKFLAG=FALSE,FRRBINDINGFLAG=FALSE,VIRTUALSYSFLAG=TRUE,TIMERCONSERFLAG=FALSE;
ADD OSPFINTERFACE:PROCID={process_id},AREAID="0.0.0.10",IFNAME="Eth-trunk2.{upftls_vlan}",NETWORKTYPE=broadcast,HOLDDOWNITVL=10,FRRBLOCKFLAG=FALSE,FRRBINDINGFLAG=FALSE,VIRTUALSYSFLAG=TRUE,TIMERCONSERFLAG=FALSE;
ADD OSPFIMPORTROUTE:PROCID={process_id},TOPOID=0,PROTOCOL=wlr,IMPTCOSTCFG=FALSE,IMPTTAGCFG=FALSE,IMPTTYPECFG=FALSE;

==== Command to Check BGP/OSPF =====
DSP OSPFPEER:;
DSP BGPPEERINFO: VRFNAME="{vrf_name}";


===================Rollback=====================================
RMV OSPFIMPORTROUTE: PROCID={process_id}, PROTOCOL=wlr;
RMV OSPFINTERFACE: PROCID={process_id}, AREAID="0.0.0.10", IFNAME="Eth-trunk1.{upftls_vlan}";
RMV OSPFINTERFACE: PROCID={process_id}, AREAID="0.0.0.10", IFNAME="Eth-trunk2.{upftls_vlan}";
RMV OSPFAREA: PROCID={process_id}, AREAID="0.0.0.10";
RMV OSPF: PROCID={process_id};
RMV IFIPV4ADDRESS: IFNAME="Eth-trunk1.{upftls_vlan}", IFIPADDR="{upftls_ip_trunk1}";
RMV IFIPV4ADDRESS: IFNAME="Eth-trunk2.{upftls_vlan}", IFIPADDR="{upftls_ip_trunk2}";
RMV ETHSUBIF: IFNAME="Eth-trunk1.{upftls_vlan}";
RMV ETHSUBIF: IFNAME="Eth-trunk2.{upftls_vlan}";
RMV IPBINDVPN: IFNAME="Eth-trunk1.{upftls_vlan}", VRFNAME="{vrf_name}";
RMV IPBINDVPN: IFNAME="Eth-trunk2.{upftls_vlan}", VRFNAME="{vrf_name}";
RMV INTERFACE: IFNAME="Eth-trunk1.{upftls_vlan}";
RMV INTERFACE: IFNAME="Eth-trunk2.{upftls_vlan}";
RMV VPNINSTAF: VRFNAME="{vrf_name}", AFTYPE=ipv4uni;
RMV L3VPNINST: VRFNAME="{vrf_name}";


            '''.format(vrf_name=vrf_name,process_id=process_id,upftls_routerid=upftls_routerid,upftls_ip_trunk1=upftls_ip_trunk1,
                       upftls_ip_trunk2=upftls_ip_trunk2,upftls_vlan=upftls_vlan)

                    created_file = datetime.today().strftime("%Y%m%d")

                    file_name = f'./TLS/UPF/UPFTL21H-{vrf_name}-{created_file}.txt'

                    with open(file_name, "w+") as f:
                        f.write(UPFTL21H)
                with col2:
                    container = st.container(border=True)
                    with open(file_name,'r') as file:
                        for text in file:
                            container.text(text)  #.text does not do any formatting
                            time.sleep(0.05)



                UPF_files = glob.glob('TLS/UPF/*')
                LATEST_UPF_FILE = max(UPF_files,key=os.path.getmtime)

                NFVI_files = glob.glob('TLS/NFVI/*')
                LATEST_NFVI_FILE = max(NFVI_files, key=os.path.getmtime)

                with open(LATEST_NFVI_FILE, 'rb') as f_nvfi:
                    st.download_button(f'Download: {LATEST_NFVI_FILE}', f_nvfi, file_name=f'{LATEST_NFVI_FILE}')

                with open(LATEST_UPF_FILE, 'rb') as f_upf:
                    st.download_button(f'Download: {LATEST_UPF_FILE}', f_upf, file_name=f'{LATEST_UPF_FILE}')
########################################################################################################################
            elif lld_file_name[2] == 'DCSLA4H':
                wb_sla = SLA_Workbook(vpn_name,wb)
                sh_routing = wb_sla.workbook['Routing']  # 19 sheet:routing
                sh_ipassingment = wb_sla.workbook['IP Assignment']  # 10 sheet:IIP Assignment
                sh_intIp = wb_sla.workbook['Internal IP Assignment']  # 12 sheet:internal IP Assignment
                ip = wb_sla.get_ipAssingment(vpn_name, sh_ipassingment)
                ipInt = wb_sla.get_intIp(vpn_name, sh_intIp)
                routing = wb_sla.get_route(sh_routing, vpn_name)

                ############ NFVISLA4H_DCGW01 ###########################################
                vnf_dcgw01 = routing[0][0:15]  # split list of vnf design
                vnf_dcgw02 = routing[0][16:24]  # split list of vnf design
                bfd1arm_dcgw01 = routing[1][0:11]
                bfd1arm_dcgw02 = routing[1][12:17]
                static_rt_dcgw01 = routing[1][17:28]
                static_rt_dcgw02 = routing[1][29:37]
                bfd_dcgw01 = routing[1][37:60]
                bfd_dcgw02 = routing[1][61:83]
                vnf_design_dcgw01 = routing[1][83:90]
                vnf_design_dcgw02 = routing[1][90:97]
                bgp_design = routing[1][97:]
                vrf_name = ip[0][6]
                corp_name = vrf_name.split('_')[-1]
                l3vni = vnf_dcgw01[7]
                as_number = bfd1arm_dcgw01[9]
                as_number_ipcbb = bgp_design[9]
                dcgw01_vlan = ip[0][8]
                dcgw01_wan_ip = ip[0][7]
                dcgw01_wan_ip = dcgw01_wan_ip.replace('/30', '')
                ipcbb01_ip = dcgw01_wan_ip.split(".")
                last_digit = int(ipcbb01_ip[3]) - 1
                ipcbb01_ip[3] = str(last_digit)
                ipcbb01_ip = '.'.join(ipcbb01_ip)
                dcgw01_loopback = bfd1arm_dcgw01[7].lower()
                dcgw01_loopback = dcgw01_loopback.replace('loopback', '')
                dcgw01_loopbackIp = bfd1arm_dcgw01[8]
                upfsla_ip = bfd1arm_dcgw01[3]
                dcgw01_l3int = vnf_dcgw01[4]
                dcgw01_bfd_ip = bfd_dcgw01[17]
                upfsla01_bfd_ip = bfd_dcgw01[8]
                dcgw01_discriminator_no = bfd_dcgw01[20]
                if len(bgp_design) == 22:
                    authentication_key = "peer {} password cipher {}".format(ipcbb01_ip, bgp_design[13])
                else:
                    authentication_key = "#No BGP authentication key"
                ############ NFVISLA4H_DCGW02 ###########################################
                # dcgw02_vlan = ip[1][8]
                # dcgw02_wan_ip = ip[1][7]
                # dcgw02_wan_ip = dcgw02_wan_ip.replace('/30','')
                # ipcbb02_ip = dcgw02_wan_ip.split(".")
                # ipcbb02_ip_lastdigit = int(ipcbb02_ip[3])-1
                # ipcbb02_ip[3] = str(ipcbb02_ip_lastdigit)
                # ipcbb02_ip = '.'.join(ipcbb02_ip)
                # dcgw02_loopback = bfd1arm_dcgw02[1].lower()
                # dcgw02_loopback = dcgw02_loopback.replace('loopback','')
                # dcgw02_loopbackIp = bfd1arm_dcgw02[2]

                NFVISLA4H_DCGW01 = ('''
/////NFVISLA4H_DCGW01 (NFVISLA4H_RN06_DCGW01)/////
/////Allow VLAN 
#
vlan batch {dcgw01_vlan}
#
interface Eth-Trunk115
 port trunk allow-pass vlan {dcgw01_vlan}
#
/////Create VPN instance
#
ip vpn-instance {vrf_name}
 description {vrf_name}
 ipv4-family
  route-distinguisher {dcgw01_wan_ip}:100
  apply-label per-instance
  vpn-target 4201000606:{l3vni} export-extcommunity evpn
  vpn-target 4201000606:{l3vni} import-extcommunity evpn
  vxlan vni {l3vni}
#
/////Configure WAN integrate with IPCBB
#
interface Vlanif{dcgw01_vlan}
 description IP_WAN_N6_IPCBB_{corp_name}_to_IPCBB_PE01
 ip binding vpn-instance {vrf_name}
 ip address {dcgw01_wan_ip} 255.255.255.252
#
///Once you have reached this step You can test ping by following command.///
ping -vpn {vrf_name} -a {dcgw01_wan_ip} {ipcbb01_ip}

/////Create interface loopback at DCGW for BGP (to UPF)
#
interface LoopBack{dcgw01_loopback}
 description UPFSLA1H {vrf_name}
 ip binding vpn-instance {vrf_name}
 ip address {dcgw01_loopbackIp} 255.255.255.255
#
/////BGP to IPCBB
#
bgp 4201000606
ipv4-family vpn-instance {vrf_name}
 as-number {as_number}
 router-id {dcgw01_wan_ip}
 network 0.0.0.0
 network {dcgw01_loopbackIp} 255.255.255.255
 import-route direct
 import-route static
 auto-frr
 advertise l2vpn evpn import-route-multipath
 peer {ipcbb01_ip} as-number {as_number_ipcbb}
 peer {ipcbb01_ip} connect-interface Vlanif{dcgw01_vlan}
 {authentication_key}
 peer {ipcbb01_ip} ip-prefix block_internal export
 peer {upfsla_ip} as-number 4201000603
 peer {upfsla_ip} description {vrf_name} to UPFSLA1H
 peer {upfsla_ip} ebgp-max-hop 10
 peer {upfsla_ip} connect-interface LoopBack{dcgw01_loopback}	
#
/////### create EVPN instances and NVE interfaces ###
evpn
#
evpn vpn-instance evpn{dcgw01_l3int} bd-mode
 route-distinguisher 10.210.205.32:{dcgw01_l3int}
 vpn-target 4201000606:{dcgw01_l3int} export-extcommunity
 vpn-target 4201000606:{dcgw01_l3int} import-extcommunity
#
interface nve1
 vni {dcgw01_l3int} head-end peer-list protocol bgp
#
bridge-domain {dcgw01_l3int}
 vxlan vni {dcgw01_l3int} split-horizon-mode
 evpn binding vpn-instance evpn{dcgw01_l3int}
#
interface Vbdif{dcgw01_l3int}
 mtu 4000
 ip binding vpn-instance {vrf_name}
 ip address {dcgw01_bfd_ip} 255.255.255.248
 arp generate-rd-table enable
 arp broadcast-detect enable
 vxlan anycast-gateway enable
 arp collect host enable
#
interface Eth-Trunk3.{dcgw01_l3int} mode l2
 encapsulation dot1q vid {dcgw01_l3int}
 rewrite pop single
 bridge-domain {dcgw01_l3int}
#
interface Eth-Trunk4.{dcgw01_l3int} mode l2
 encapsulation dot1q vid {dcgw01_l3int}
 rewrite pop single
 bridge-domain {dcgw01_l3int}
#
interface Eth-Trunk71.{dcgw01_l3int} mode l2
 encapsulation dot1q vid {dcgw01_l3int}
 rewrite pop single
 bridge-domain {dcgw01_l3int}
#
/////BFD to UPF
#
bfd N6_IPCBB_{corp_name}_UPFSLA1H_01 bind peer-ip {upfsla01_bfd_ip} vpn-instance {vrf_name} interface Vbdif{dcgw01_l3int} one-arm-echo destination-ip {dcgw01_loopbackIp}
 discriminator local {dcgw01_discriminator_no}
 detect-multiplier 4
 wtr 3
 min-echo-rx-interval 500
#
/////Static Route to UPF (Tag BFD)
#
ip route-static vpn-instance {vrf_name} {upfsla_ip} 255.255.255.255 {upfsla01_bfd_ip} track bfd-session N6_IPCBB_{corp_name}_UPFSLA1H_01 description Loopback {vrf_name} of UPFSLA1H
# 
commit
save 

==== Command to Check eBGP state to IPCBB (Shall be Established) =====
display bgp vpnv4 vpn-instance {vrf_name} peer


///////Roll back////////////
///NFVISLA4H_DCGW01///
#
undo ip route-static vpn-instance {vrf_name} {upfsla_ip} 255.255.255.255 {upfsla01_bfd_ip}
undo bfd N6_IPCBB_{corp_name}_UPFSLA1H_01
#
bgp 4201000606
undo ipv4-family vpn-instance {vrf_name}
#
undo interface Eth-Trunk3.{dcgw01_l3int}
undo interface Eth-Trunk4.{dcgw01_l3int}
undo interface Eth-Trunk71.{dcgw01_l3int}
#
undo interface Vlanif{dcgw01_vlan}
undo interface LoopBack{dcgw01_loopback}
undo interface Vbdif{dcgw01_l3int}
#
interface nve1
undo  vni {dcgw01_l3int}
#
undo bridge-domain {dcgw01_l3int}
#
undo evpn vpn-instance evpn{dcgw01_l3int} bd-mode
#
interface Eth-Trunk115
undo  port trunk allow-pass vlan {dcgw01_vlan}
#
undo ip vpn-instance {vrf_name}
#
undo vlan batch {dcgw01_vlan}

commit
return
save

# '''.format(dcgw01_vlan=dcgw01_vlan, vrf_name=vrf_name, corp_name=corp_name,
                   dcgw01_wan_ip=dcgw01_wan_ip, ipcbb01_ip=ipcbb01_ip,
                   l3vni=l3vni, dcgw01_loopback=dcgw01_loopback, dcgw01_loopbackIp=dcgw01_loopbackIp,
                   as_number=as_number, upfsla_ip=upfsla_ip,
                   dcgw01_l3int=dcgw01_l3int, as_number_ipcbb=as_number_ipcbb, dcgw01_bfd_ip=dcgw01_bfd_ip,
                   upfsla01_bfd_ip=upfsla01_bfd_ip,
                   dcgw01_discriminator_no=dcgw01_discriminator_no, authentication_key=authentication_key))

                col1, col2 = st.columns(2)
                created_file = datetime.today().strftime("%Y%m%d")
                file_name = f'./SLA/NFVI/DCGW01/NFVISLA4H-DCGW01-{vrf_name}-{created_file}.txt'

                with open(file_name, "w+") as f:
                    f.write(NFVISLA4H_DCGW01)
                with col1:
                    container = st.container(border=True)
                    with open(file_name, 'r') as file:
                        for text in file:
                            container.text(text)  # .text does not do any formatting
                            time.sleep(0.05)

                ############ NFVISLA4H_DCGW02 ###########################################
                dcgw02_vlan = ip[1][8]
                dcgw02_wan_ip = ip[1][7]
                dcgw02_wan_ip = dcgw02_wan_ip.replace('/30', '')
                ipcbb02_ip = dcgw02_wan_ip.split(".")
                ipcbb02_ip_lastdigit = int(ipcbb02_ip[3]) - 1
                ipcbb02_ip[3] = str(ipcbb02_ip_lastdigit)
                ipcbb02_ip = '.'.join(ipcbb02_ip)
                dcgw02_loopback = bfd1arm_dcgw02[1].lower()
                dcgw02_loopback = dcgw02_loopback.replace('loopback', '')
                dcgw02_loopbackIp = bfd1arm_dcgw02[2]
                dcgw02_l3int = vnf_dcgw02[3]
                dcgw02_bfd_ip = bfd_dcgw02[16]
                upfsla02_bfd_ip = bfd_dcgw02[7]
                dcgw02_discriminator_no = bfd_dcgw02[19]
                if len(bgp_design) == 22:
                    authentication_key = "peer {} password cipher {}".format(ipcbb02_ip, bgp_design[13])
                else:
                    authentication_key = "#No BGP authentication key"

                NFVISLA4H_DCGW02 = ('''
/////NFVISLA4H_DCGW02 (NFVISLA4H_RN07_DCGW01)/////
/////Allow VLAN 
#
vlan batch {dcgw02_vlan}
#
interface Eth-Trunk116
 port trunk allow-pass vlan {dcgw02_vlan}
#
/////Create VPN instance
#
ip vpn-instance {vrf_name}
 description {vrf_name}
 ipv4-family
  route-distinguisher {dcgw02_wan_ip}:100
  apply-label per-instance
  vpn-target 4201000606:{l3vni} export-extcommunity evpn
  vpn-target 4201000606:{l3vni} import-extcommunity evpn
 vxlan vni {l3vni}
#
/////Configure WAN integrate with IPCBB
#
interface Vlanif{dcgw02_vlan}
 description IP_WAN_N6_IPCBB_{corp_name}_to_IPCBB_PE02
 ip binding vpn-instance {vrf_name}
 ip address {dcgw02_wan_ip} 255.255.255.252
#
///Once you have reached this step You can test ping by following command.///
ping -vpn {vrf_name} -a {dcgw02_wan_ip} {ipcbb02_ip}

/////Create interface loopback at DCGW for BGP (to UPF)
#
interface LoopBack{dcgw02_loopback}
 description UPFSLA1H {vrf_name}
 ip binding vpn-instance {vrf_name}
 ip address {dcgw02_loopbackIp} 255.255.255.255
#
/////BGP to IPCBB
#
bgp 4201000606
ipv4-family vpn-instance {vrf_name}
 as-number {as_number}
 router-id {dcgw02_wan_ip}
 network 0.0.0.0
 network {dcgw02_loopbackIp} 255.255.255.255
 import-route direct
 import-route static
 auto-frr
 advertise l2vpn evpn import-route-multipath
 peer {ipcbb02_ip} as-number {as_number_ipcbb}
 peer {ipcbb02_ip} connect-interface Vlanif{dcgw02_vlan}
 {authentication_key}
 peer {ipcbb02_ip} ip-prefix block_internal export
 peer {upfsla_ip} as-number 4201000603
 peer {upfsla_ip} description {vrf_name} to UPFSLA1H
 peer {upfsla_ip} ebgp-max-hop 10
 peer {upfsla_ip} connect-interface LoopBack{dcgw02_loopback}
#
/////### create EVPN instances and NVE interfaces ###
#
evpn vpn-instance evpn{dcgw02_l3int} bd-mode
 route-distinguisher 10.210.205.33:{dcgw02_l3int}
 vpn-target 4201000606:{dcgw02_l3int} export-extcommunity
 vpn-target 4201000606:{dcgw02_l3int} import-extcommunity
#
interface nve1
 vni {dcgw02_l3int} head-end peer-list protocol bgp
#
bridge-domain {dcgw02_l3int}
 vxlan vni {dcgw02_l3int} split-horizon-mode
 evpn binding vpn-instance evpn{dcgw02_l3int}
#
interface Vbdif{dcgw02_l3int}
 mtu 4000
 ip binding vpn-instance {vrf_name}
 ip address {dcgw02_bfd_ip} 255.255.255.248
 arp generate-rd-table enable
 arp broadcast-detect enable
 vxlan anycast-gateway enable
 arp collect host enable
#
interface Eth-Trunk3.{dcgw02_l3int} mode l2
 encapsulation dot1q vid {dcgw02_l3int}
 rewrite pop single
 bridge-domain {dcgw02_l3int}
#
interface Eth-Trunk4.{dcgw02_l3int} mode l2
 encapsulation dot1q vid {dcgw02_l3int}
 rewrite pop single
 bridge-domain {dcgw02_l3int}
#
interface Eth-Trunk72.{dcgw02_l3int} mode l2
 encapsulation dot1q vid {dcgw02_l3int}
 rewrite pop single
 bridge-domain {dcgw02_l3int}
#
/////BFD to UPF
#
bfd N6_IPCBB_{corp_name}_UPFSLA1H_02 bind peer-ip {upfsla02_bfd_ip} vpn-instance {vrf_name} interface Vbdif{dcgw02_l3int} one-arm-echo destination-ip {dcgw02_loopbackIp}
 discriminator local {dcgw02_discriminator_no}
 detect-multiplier 4
 wtr 3
 min-echo-rx-interval 500
#
/////Static Route to UPF (Tag BFD)
#
ip route-static vpn-instance {vrf_name} {upfsla_ip} 255.255.255.255 {upfsla02_bfd_ip} track bfd-session N6_IPCBB_{corp_name}_UPFSLA1H_02 description Loopback {vrf_name} of UPFSLA1H
#
commit
save 

==== Command to Check eBGP state to IPCBB (Shall be Established) =====
display bgp vpnv4 vpn-instance {vrf_name} peer 


//////////////Roll back///////////////////////
///NFVISLA4H_DCGW02///
#
undo ip route-static vpn-instance {vrf_name} {upfsla_ip} 255.255.255.255 {upfsla02_bfd_ip}
undo bfd N6_IPCBB_{corp_name}_UPFSLA1H_02

#
bgp 4201000606
undo ipv4-family vpn-instance {vrf_name}
#
undo interface Eth-Trunk3.{dcgw02_l3int}
undo interface Eth-Trunk4.{dcgw02_l3int}
undo interface Eth-Trunk72.{dcgw02_l3int}
#
undo interface Vlanif{dcgw02_vlan}
undo interface LoopBack{dcgw02_loopback}
undo interface Vbdif{dcgw02_l3int}
#
interface nve1
undo  vni {dcgw02_l3int}
#
undo bridge-domain {dcgw02_l3int}
#
undo evpn vpn-instance evpn{dcgw02_l3int} bd-mode
#
interface Eth-Trunk116
undo  port trunk allow-pass vlan {dcgw02_vlan}
#
undo ip vpn-instance {vrf_name}
#
undo vlan batch {dcgw02_vlan}

commit
return
save

                '''.format(dcgw02_vlan=dcgw02_vlan, vrf_name=vrf_name, corp_name=corp_name,
                           dcgw02_wan_ip=dcgw02_wan_ip, ipcbb02_ip=ipcbb02_ip,
                           l3vni=l3vni, dcgw02_loopback=dcgw02_loopback, dcgw02_loopbackIp=dcgw02_loopbackIp,
                           as_number=as_number, upfsla_ip=upfsla_ip,
                           dcgw02_l3int=dcgw02_l3int, as_number_ipcbb=as_number_ipcbb, dcgw02_bfd_ip=dcgw02_bfd_ip,
                           upfsla02_bfd_ip=upfsla02_bfd_ip,
                           dcgw02_discriminator_no=dcgw02_discriminator_no, authentication_key=authentication_key))

                # print(sla_DCGW02)
                created_file = datetime.today().strftime("%Y%m%d")
                file_name = f'./SLA/NFVI/DCGW02/NFVISLA4H_DCGW02-{vrf_name}-{created_file}.txt'
                with open(file_name, "w+") as f:
                    f.write(NFVISLA4H_DCGW02)
                with col2:
                    container = st.container(border=True)
                    with open(file_name, 'r') as file:
                        for text in file:
                            container.text(text)  # .text does not do any formatting
                            time.sleep(0.05)


                ############ UPFSLA1H ###########################################
                upfsla_loopback = bfd1arm_dcgw01[2].lower()
                upfsla_loopback = upfsla_loopback.replace('loopback', '')
                vrfrd = vnf_dcgw01[2]
                upfsla01_intIp = ipInt[0][4]
                upfsla02_intIp = ipInt[1][4]
                ipcbb01_intIp = ipInt[0][7]
                ipcbb02_intIp = ipInt[1][7]
                route_id = bfd1arm_dcgw01[1]
                upfsla01_discriminator_no = bfd_dcgw01[11]
                upfsla02_discriminator_no = bfd_dcgw02[10]

                UPFSLA1H = '''
ADD L3VPNINST:VRFNAME="{vrf_name}";
ADD VPNINSTAF:VRFNAME="{vrf_name}",AFTYPE=ipv4uni,VRFRD="{vrfrd}";
ADD INTERFACE:IFNAME="Eth-trunk1.{dcgw01_l3int}";
ADD INTERFACE:IFNAME="Eth-trunk2.{dcgw02_l3int}";
ADD INTERFACE:IFNAME="LoopBack{upfsla_loopback}";
ADD IPBINDVPN:IFNAME="Eth-trunk1.{dcgw01_l3int}",VRFNAME="{vrf_name}";
ADD IPBINDVPN:IFNAME="Eth-trunk2.{dcgw02_l3int}",VRFNAME="{vrf_name}";
ADD IPBINDVPN:IFNAME="LoopBack{upfsla_loopback}",VRFNAME="{vrf_name}";
ADD ETHSUBIF:IFNAME="Eth-trunk1.{dcgw01_l3int}",VLANTYPEVID={dcgw01_l3int};
ADD ETHSUBIF:IFNAME="Eth-trunk2.{dcgw02_l3int}",VLANTYPEVID={dcgw02_l3int};
ADD IFIPV4ADDRESS:IFNAME="Eth-trunk1.{dcgw01_l3int}",IFIPADDR="{upfsla01_intIp}",SUBNETMASK="255.255.255.248";
ADD IFIPV4ADDRESS:IFNAME="Eth-trunk2.{dcgw02_l3int}",IFIPADDR="{upfsla02_intIp}",SUBNETMASK="255.255.255.248";
ADD IFIPV4ADDRESS:IFNAME="LoopBack{upfsla_loopback}",IFIPADDR="{upfsla_ip}",SUBNETMASK="255.255.255.255";
ADD BFDSESSION: SESSNAME="BFD_N6_IPCBB_{corp_name}_01", ADDRTYPE=IPv4, CREATETYPE4=SESS_STATIC, ONEARMECHO=TRUE, DESTADDR4="{ipcbb01_intIp}", LINKTYPE=IP, VRFNAME="{vrf_name}", LOCALDISCR={upfsla01_discriminator_no}, IFNAME="Eth-trunk1.{dcgw01_l3int}";
ADD BFDSESSION: SESSNAME="BFD_N6_IPCBB_{corp_name}_02", ADDRTYPE=IPv4, CREATETYPE4=SESS_STATIC, ONEARMECHO=TRUE, DESTADDR4="{ipcbb02_intIp}", LINKTYPE=IP, VRFNAME="{vrf_name}", LOCALDISCR={upfsla02_discriminator_no}, IFNAME="Eth-trunk2.{dcgw02_l3int}";
ADD SRROUTE: AFTYPE=ipv4unicast, PREFIX="0.0.0.0", MASKLENGTH=0, VRFNAME="{vrf_name}", DESTVRFNAME="{vrf_name}", IFNAME="Eth-trunk1.{dcgw01_l3int}", NEXTHOP="{ipcbb01_intIp}", SESSIONNAME="BFD_N6_IPCBB_{corp_name}_01";
ADD SRROUTE: AFTYPE=ipv4unicast, PREFIX="0.0.0.0", MASKLENGTH=0, VRFNAME="{vrf_name}", DESTVRFNAME="{vrf_name}", IFNAME="Eth-trunk2.{dcgw02_l3int}", NEXTHOP="{ipcbb02_intIp}", SESSIONNAME="BFD_N6_IPCBB_{corp_name}_02";
ADD BGPVRF:VRFNAME="{vrf_name}",DEFAULTAFTYPE=noaf;
ADD BGPVRFAF:VRFNAME="{vrf_name}",AFTYPE=ipv4uni,MAXIMUMLOADBALANCE=2,ROUTERID="{route_id}";
ADD BGPPEER:VRFNAME="{vrf_name}",ADDRESSTYPE=ipv4,PEERADDR="{dcgw01_loopbackIp}",REMOTEAS="{as_number}",LOCALIFADDR="{upfsla_ip}",LOCALIFNAME="LoopBack{upfsla_loopback}",EBGPMAXHOP=10;
ADD BGPPEER:VRFNAME="{vrf_name}",ADDRESSTYPE=ipv4,PEERADDR="{dcgw02_loopbackIp}",REMOTEAS="{as_number}",LOCALIFADDR="{upfsla_ip}",LOCALIFNAME="LoopBack{upfsla_loopback}",EBGPMAXHOP=10;
ADD IMPORTROUTE:VRFNAME="{vrf_name}",AFTYPE=ipv4uni,IMPORTPROTOCOL=wlr,MEDENABLE=FALSE;

==== Command to Check BGP/BFD =====
DSP BGPPEERINFO: VRFNAME="{vrf_name}";
DSP BFDSESSION:;


//rollback
RMV INTERFACE:IFNAME="Eth-trunk1.{dcgw01_l3int}";
RMV INTERFACE:IFNAME="Eth-trunk2.{dcgw02_l3int}";
RMV INTERFACE:IFNAME="LoopBack{upfsla_loopback}";


                    '''.format(vrf_name=vrf_name, corp_name=corp_name, dcgw01_l3int=dcgw01_l3int,
                               dcgw02_l3int=dcgw02_l3int, as_number=as_number,
                               upfsla_loopback=upfsla_loopback, vrfrd=vrfrd, upfsla01_intIp=upfsla01_intIp,
                               upfsla02_intIp=upfsla02_intIp,
                               upfsla_ip=upfsla_ip, ipcbb01_intIp=ipcbb01_intIp, ipcbb02_intIp=ipcbb02_intIp,
                               upfsla01_discriminator_no=upfsla01_discriminator_no,
                               upfsla02_discriminator_no=upfsla02_discriminator_no, route_id=route_id,
                               dcgw01_loopbackIp=dcgw01_loopbackIp, dcgw02_loopbackIp=dcgw02_loopbackIp)


                created_file = datetime.today().strftime("%Y%m%d")
                file_name = f'./SLA/UPF/UPFSLA1H-{vrf_name}-{created_file}.txt'

                with open(file_name, "w+") as f:
                    f.write(UPFSLA1H)

                container = st.container(border=True)
                with open(file_name,'r') as file:
                    for text in file:
                        container.text(text)  #.text does not do any formatting
                        time.sleep(0.05)


                # print(UPFSLA1H)
                UPF_files = glob.glob('SLA/UPF/*')
                LATEST_UPF_FILE = max(UPF_files, key=os.path.getmtime)

                NFVI01_files = glob.glob('SLA/NFVI/DCGW01/*')
                LATEST_NFVI01_FILE = max(NFVI01_files, key=os.path.getmtime)

                NFVI02_files = glob.glob('SLA/NFVI/DCGW02/*')
                LATEST_NFVI02_FILE = max(NFVI02_files, key=os.path.getmtime)

                with open(LATEST_NFVI01_FILE, 'rb') as f_nvfi:
                    st.download_button(f'Download: {LATEST_NFVI01_FILE}', f_nvfi, file_name=f'{LATEST_NFVI01_FILE}')

                with open(LATEST_NFVI02_FILE, 'rb') as f_nvfi:
                    st.download_button(f'Download: {LATEST_NFVI02_FILE}', f_nvfi, file_name=f'{LATEST_NFVI02_FILE}')

                with open(LATEST_UPF_FILE, 'rb') as f_upf:
                    st.download_button(f'Download: {LATEST_UPF_FILE}', f_upf, file_name=f'{LATEST_UPF_FILE}')


        except:
                st.write('Check loaded file or VPN name in LLD is correct?')
    except:
        st.write()


