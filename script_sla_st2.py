import openpyxl as xl
import pandas as pd

# vpnName = "VPN_N6_IPCBB_TOA"
# lld_sla ="CloudDesign LLD DCSLA4H (20240223).xlsx"
# df_ipAssignment = pd.read_excel(lld_sla,sheet_name='IP Assignment',skiprows=[0])
# df_internalIpAssignment = pd.read_excel(lld_sla,sheet_name='Internal IP Assignment')
# df_route = pd.read_excel(lld_sla,sheet_name='Routing',skiprows=[0,17])
# ws = xl.load_workbook(lld_sla,data_only=True)
# sh_routing = ws.worksheets[19]  # 19 sheet:routing
# #NFVI to IPCBB

class SLA_Workbook:
    def __init__(self,vpn,workbook):
        self.vpn = vpn
        self.workbook = workbook


    def get_ipAssingment(self,vpn,sheet):
        i = 0
        list_rowid = []
        list_ip_dcgw01 = []
        list_ip_dcgw02 = []
        for c in sheet.values:
            i += 1
            if c[6] == vpn:
                list_rowid.append(i)

        for r in sheet.iter_rows(min_row=list_rowid[0], max_row=list_rowid[0]):
            for v in r:
                list_ip_dcgw01.append(v.value)

        for r in sheet.iter_rows(min_row=list_rowid[1], max_row=list_rowid[1]):
            for v in r:
                list_ip_dcgw02.append(v.value)
        return list_ip_dcgw01,list_ip_dcgw02

    def get_intIp(self,vpn,sheet):
        i = 0
        list_rowid = []
        list_intIp_dcgw01 = []
        list_intIp_dcgw02 = []
        for c in sheet.values:
            i += 1
            if c[2] == vpn:
                list_rowid.append(i)

        for r in sheet.iter_rows(min_row=list_rowid[0], max_row=list_rowid[0]):
            for v in r:
                list_intIp_dcgw01.append(v.value)

        for r in sheet.iter_rows(min_row=list_rowid[1], max_row=list_rowid[1]):
            for v in r:
                list_intIp_dcgw02.append(v.value)
        return list_intIp_dcgw01,list_intIp_dcgw02



    def get_route(self,sheet,vpn):
          #get value for formular cell
          list1_vpn_rowid = []
          list2_vpn_rowid = []
          list3_vpn_rowid = []
          i = 0
          for v in sheet.values:
              i += 1
              if v[1] == vpn:
                  list1_vpn_rowid.append(i)

              if v[2] == vpn:
                  z = i + 1
                  list3_vpn_rowid.append(i)
                  list3_vpn_rowid.insert(1, z)
          # print(list3_vpn_rowid)

          for x in list1_vpn_rowid:
              n = x + 1
              list2_vpn_rowid.append(n)

          list1_vpn_rowid.insert(1, list2_vpn_rowid[0])
          list1_vpn_rowid.insert(3, list2_vpn_rowid[1])
          # print(list1_vpn_rowid)

          list_dcgw01_t1 = []
          for t1 in list3_vpn_rowid:
              for c in sheet.iter_rows(min_row=t1, max_row=t1):
                  for z in c:
                      if z.value != None:
                        list_dcgw01_t1.append(z.value)
          # print(list_dcgw01_t1)

          list_dcgw01 = []
          for l in list1_vpn_rowid:
              for v in sheet.iter_rows(min_row=l, max_row=l):
                  for y in v:
                      if y.value != None:
                          list_dcgw01.append(y.value)
          # print(list_dcgw01)

          return list_dcgw01_t1,list_dcgw01

